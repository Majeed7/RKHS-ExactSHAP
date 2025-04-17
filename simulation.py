import numpy as np 
import matplotlib.pyplot as plt
import time
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from sklearn.preprocessing import StandardScaler

from utils.game_simulation_utils import *
from utils.real_datasets import load_dataset
from utils.svm_training import optimize_svm_rbf
from explainer.LocalExplainer import *

import warnings
warnings.filterwarnings('ignore')


shap_alldiffs = []
shap_alldiffs_sd = []

shap_reg_alldiffs = []
shap_reg_alldiffs_sd = []

gemfix_alldiffs = []
gemfix_alldiffs_sd = []

gemfix_reg_alldiffs = []
gemfix_reg_alldiffs_sd = []

shap_time_all = []
shap_time_std_all = []

shap_reg_time_all = []
shap_reg_time_std_all = []

gemfix_time_all = []
gemfix_time_std_all = []

gemfix_reg_time_all = []
gemfix_reg_time_std_all = []

drawn_samples = []

from sklearn.datasets import load_diabetes, fetch_california_housing, load_wine, make_regression
#data = load_diabetes() # load_wine() #fetch_california_housing() #
#X = data['data']
#y = data['target'] + 0.001
#X, y = load_dataset('breast_cancer')
X, y = make_regression(n_samples=500, n_features=50, noise=0.1)

scaler_X = StandardScaler()
scaler_y = StandardScaler()

# Fit the scaler and transform X
#X = scaler_X.fit_transform(X)

# Fit the scaler and transform y
#y = scaler_y.fit_transform(y.reshape(-1, 1)).flatten()
#y = y * 2 + 5

n, d = X.shape 

svm_optimized = optimize_svm_rbf(X,y, n_trials=50)
model = svm_optimized['model']

sampleNo_tbx = 10
indices = np.random.choice(X.shape[0], size=min(sampleNo_tbx, X.shape[0]), replace=False)
X_tbx = X[indices,:]
X_tbx = [X[0,]]
'''
this function compute y for the selected feature based on the value function given
'''
def compute_y(k_vectors, samples, alpha):
    y = []
    for s in samples:
        y.append(np.prod(k_vectors[s.astype(bool),:], axis=0).T @ alpha)

    return np.array(y)    

for x in X_tbx:

    explainer = RBFLocalExplainer(model)  # use same explainer interface as for regression
    shapley_values = explainer.explain(x)

    k_vectors = np.array(explainer.compute_kernel_vectors(X[model.support_,:], x))

    samples_weights = kernel_shap_sampling(d, 1000)
    samples = np.array([t[0] for t in samples_weights])
    weights = np.array([t[1] for t in samples_weights])

    samples = np.vstack([samples, np.ones(samples[0].shape)]).astype(int)
    weights = np.hstack([weights, 1e20])

    y = compute_y(k_vectors, samples, model.dual_coef_.squeeze()) 

    shap_reg = Shapley_regression(np.array(samples, dtype='float'), y, weights)
    shap_vals = shap_reg.coef_
    print("SHAP error: ", np.linalg.norm(shap_vals - shapley_values, 1) / np.sum(shapley_values))

    weights_mobius = np.ones(samples[:,0].shape)
    weights_mobius[-1] = 1e30

    gemfix_vals = gemfix_reg(samples.astype('float'), y, weights_mobius)

    print("GEM-FIX error: ", np.linalg.norm(gemfix_vals - shapley_values, 1) / np.sum(shapley_values))

for N in range(1):
    print(f"the number of player is: {N}")
    players, subset_values_mobius, subsets_all, subset_values = generate_game(N, N, essential=False) # generate a game
    exact_sv = exact_Shapley_value(subset_values_mobius, N) # computing the exact Shapely value of the game
    #subsets_all.pop(0)
    
    max_samplesize = 100000
    sample_no = N if N < 14 else N*2
    sample_size = random.sample(range(2*N, np.min((2**N, max_samplesize))), sample_no)
    #sample_size.append(2 ** N - 3)
    sample_sizes = np.sort(sample_size)

    mat, y, weights = subset_to_matrix(N, subsets_all, subset_values)
    mat_mobius, y_mobius, weights_mobius = subset_to_matrix_mobius(N, subsets_all, subsets_all, subset_values)

    shap_sample_diff = []
    shap_sample_diff_sd = []

    shap_reg_sample_diff = []
    shap_reg_sample_diff_sd = []

    gemfix_sample_diff = []
    gemfix_sample_diff_sd = []

    gemfix_reg_sample_diff = []
    gemfix_reg_sample_diff_sd = []

    shap_time = np.zeros((len(sample_size),))
    shap_time_sd = np.zeros((len(sample_size),))

    shap_reg_time = np.zeros((len(sample_size),))
    shap_reg_time_sd = np.zeros((len(sample_size),))

    gemfix_time = np.zeros((len(sample_size),))
    gemfix_time_sd = np.zeros((len(sample_size),))

    gemfix_reg_time = np.zeros((len(sample_size),))
    gemfix_reg_time_sd = np.zeros((len(sample_size),))

    drawn_sample_N = []
    for index, n_samples in enumerate(sample_sizes):
        print(f"smaple size: {n_samples}")
        drawn_sample_N.append(n_samples)
           
        shap_rep = []
        shap_reg_rep = []

        gemfix_rep = []
        shapley_mob_rr_rep = []
        shapley_mob_lasso_rep = []
        gemfix_reg_rep = []

        shap_rep_t = []
        shap_reg_rep_t = []
        gemfix_rep_t = []
        gemfix_reg_rep_t = []

        for _ in range(replication):
            sample_index = random.sample(range(2 ** N - 1), n_samples)    
            sample_index.append(-1) # add the grand coalition        
            sample_index.append(0)
            
            ## SHAP Linear Regression approximation of Shapley value; without regularization 
            start = time.time()
            shap_val = Shapley_regression(mat[sample_index,:], y[sample_index], weights[sample_index])
            shap_rep.append(np.linalg.norm(shap_val.coef_ - exact_sv, 1))
            shap_rep_t.append(time.time() - start)
            
            ## SHAP Linear Regression approximation of Shapley value; with L2 regularization 
            start = time.time()
            shap_reg_val = Shapley_regression(mat[sample_index,:], y[sample_index], weights[sample_index], model_type='shap_reg')
            shap_reg_rep.append(np.linalg.norm(shap_reg_val.coef_ - exact_sv, 1))
            shap_reg_rep_t.append(time.time() - start)

            ## GEM-FIX linear regression             
            start = time.time()
            if N <= 12:
                gemfix_shapley, model_gemfix = gemfix_regression(mat_mobius[sample_index,:], y_mobius[sample_index], weights_mobius[sample_index], subsets_all, N)
                gemfix_rep.append(np.linalg.norm(gemfix_shapley - exact_sv, 1))
            else:
                gemfix_rep.append(-np.ones((N,)))
            gemfix_rep_t.append(time.time() - start)

            #GEM-FIX linear regression with L2 regularization
            start = time.time()
            gemfix_reg_shapley, model_gemfix_reg = gemfix_regression(mat[sample_index,:], y[sample_index], weights_mobius[sample_index], subsets_all, N, model_type='gemfix_reg')
            gemfix_reg_rep.append(np.linalg.norm(gemfix_reg_shapley - exact_sv, 1))
            gemfix_reg_rep_t.append(time.time() - start)

        ##Storing the difference to real shapley value
        shap_sample_diff.append(np.average(shap_rep))
        shap_sample_diff_sd.append(np.std(shap_rep))

        shap_reg_sample_diff.append(np.average(shap_reg_rep))
        shap_reg_sample_diff_sd.append(np.std(shap_reg_rep))

        gemfix_sample_diff.append(np.average(gemfix_rep))
        gemfix_sample_diff_sd.append(np.std(gemfix_rep))

        gemfix_reg_sample_diff.append(np.average(gemfix_reg_rep))
        gemfix_reg_sample_diff_sd.append(np.std(gemfix_reg_rep))

        ## Storing the time stat
        shap_time[index] = np.mean(shap_rep_t)
        shap_time_sd[index] = np.std(shap_rep_t)    

        shap_reg_time[index] = np.mean(shap_reg_rep_t)
        shap_reg_time_sd[index] = np.std(shap_reg_rep_t)

        gemfix_time[index] = np.mean(gemfix_rep_t)
        gemfix_time_sd[index] = np.std(gemfix_rep_t)
        
        gemfix_reg_time[index] = np.mean(gemfix_reg_rep_t)
        gemfix_reg_time_sd[index] = np.std(gemfix_reg_rep_t)
    
    drawn_samples.append(drawn_sample_N)

    #stroing the deviation from the exact Shapley value 
    shap_alldiffs.append(shap_sample_diff)
    shap_alldiffs_sd.append(shap_sample_diff_sd)

    shap_reg_alldiffs.append(shap_reg_sample_diff)
    shap_reg_alldiffs_sd.append(shap_reg_sample_diff_sd)

    gemfix_alldiffs.append(gemfix_sample_diff)
    gemfix_alldiffs_sd.append(gemfix_sample_diff_sd)

    gemfix_reg_alldiffs.append(gemfix_reg_sample_diff)
    gemfix_reg_alldiffs_sd.append(gemfix_reg_sample_diff_sd)


    ## storing the time
    shap_time_all.append(shap_time)
    shap_time_std_all.append(shap_time_sd)

    shap_reg_time_all.append(shap_reg_time)
    shap_reg_time_std_all.append(shap_reg_time_sd)

    gemfix_time_all.append(gemfix_time)
    gemfix_time_std_all.append(gemfix_time_sd)

    gemfix_reg_time_all.append(gemfix_reg_time)
    gemfix_reg_time_std_all.append(gemfix_reg_time_sd)
   
    ## Saving to Excel file
    row_header_diff = ['SHAP', 'SHAP_sd', 'SHAP-L2', 'SHAP-L2_sd', 'GEMFIX-LR', 'GEMFIX-LR_sd', 'GEMFIX', 'GEMFIX_sd']
    df = pd.DataFrame([shap_sample_diff, shap_sample_diff_sd,  shap_reg_sample_diff, shap_reg_sample_diff_sd,
                       gemfix_sample_diff, gemfix_sample_diff_sd, gemfix_reg_sample_diff, gemfix_reg_sample_diff_sd], 
                       index=row_header_diff, columns=drawn_sample_N)
    
    mode = 'a' if excel_path_results.exists() else 'w'
    
    with pd.ExcelWriter(excel_path_results, engine='openpyxl', mode=mode) as writer:
        # Attempt to load the workbook if it exists to check for sheet names
        if mode == 'a':
            writer.book = load_workbook(excel_path_results)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        sheet_name = f'#players_{N}'

        counter = 1
        while sheet_name in writer.sheets:
            sheet_name = f"{sheet_name}_{(counter)}"
            counter += 1
        
        df.to_excel(writer, sheet_name=sheet_name, index_label='Method')

    
    mode_time = 'a' if excel_path_time.exists() else 'w'
    df_time = pd.DataFrame([shap_time, shap_time_sd,  shap_reg_time, shap_reg_time_sd,
            gemfix_time, gemfix_time_sd, gemfix_reg_time, gemfix_reg_time_sd], 
            index=row_header_diff, columns=drawn_sample_N)
    
    with pd.ExcelWriter(excel_path_time, engine='openpyxl', mode=mode_time) as writer:
        # Attempt to load the workbook if it exists to check for sheet names
        if mode == 'a':
            writer.book = load_workbook(excel_path_results)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        sheet_name = f'#players_{N}'

        counter = 1
        while sheet_name in writer.sheets:
            sheet_name = f"{sheet_name}_{(counter)}"
            counter += 1
        
        df_time.to_excel(writer, sheet_name=sheet_name, index_label='Method')

#plt.figure(figsize=(15, 10))
fig, axs = plt.subplots(len(player_range))

for i in range(len(player_range)):
    axs[i].plot(drawn_samples[i], shap_alldiffs[i], label='SHAP')
    axs[i].plot(drawn_samples[i], shap_reg_alldiffs[i], label='SHAP-L2')

    axs[i].plot(drawn_samples[i], gemfix_alldiffs[i], label='GEM-FIX-LR')
    axs[i].plot(drawn_samples[i], gemfix_reg_alldiffs[i], label='GEM-FIX')
    #axs[i].plot(drawn_samples[i], shapley_mobius_lasso_alldiffs[i], label='Mobius-lasso')

plt.xlabel('Sample Size')
plt.ylabel('Average Absolute Difference')
plt.title('Average Absolute Difference vs. Sample Size for Different Number of Features')
plt.legend()
plt.show()


fig_time, axs_time = plt.subplots((len(player_range)))
for i in range(len(player_range)):
    axs_time[i].plot(drawn_samples[i], shap_time_all[i], label='SHAP')
    axs_time[i].plot(drawn_samples[i], shap_reg_time_all[i], label='SHAP-L2')

    axs_time[i].plot(drawn_samples[i], gemfix_time_all[i], label='GEM-FIX-LR')
    axs_time[i].plot(drawn_samples[i], gemfix_reg_time_all[i], label='GEM-FIX')

plt.xlabel('Sample Size')
plt.ylabel('Time (S)')
plt.title('Execution time of methods for different number of players')
plt.legend()
plt.show()

print("done")




            
        

