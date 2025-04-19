import numpy as np 
import matplotlib.pyplot as plt
import time
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from sklearn.preprocessing import StandardScaler
from sklearn.datasets import make_regression
import os

from utils.game_simulation_utils import *
from utils.real_datasets import load_dataset
from utils.svm_training import optimize_svm_rbf
from explainer.LocalExplainer import *


plt.rcParams.update({
    'font.size': 10,     # Set default font size
    'font.weight': 'bold',  # Set default font weight to bold
    'axes.labelweight': 'bold',  # Ensure the axis labels are bold
    'axes.titleweight': 'bold',  # Ensure the titles are bold
    'figure.titleweight': 'bold',  # Bold for suptitle if you use fig.suptitle()
    'xtick.labelsize': 8,  # Font size for X-tick labels
    'ytick.labelsize': 12,  # Font size for Y-tick labels
    'xtick.major.size': 5,  # Length of major ticks
    'ytick.major.size': 5,  # Length of major ticks
    'xtick.minor.size': 3,  # Length of minor ticks
    'ytick.minor.size': 3   # Length of minor ticks
})


import warnings
warnings.filterwarnings('ignore')

def compute_y(k_vectors, samples, alpha):
    y = []
    for s in samples:
        y.append(np.prod(k_vectors[s.astype(bool),:], axis=0).T @ alpha)

    return np.array(y)  

# Define the number of features for each dataset
ds = [100] #[10, 20, 50, 100]

# Create results directory if it doesn't exist
results_dir = "results"
os.makedirs(results_dir, exist_ok=True)

if True:
    for d in ds:
        # Generate dataset
        X, y = make_regression(n_samples=50*d, n_features=d, noise=0.1)
        scaler_X = StandardScaler()
        scaler_y = StandardScaler()
        X = scaler_X.fit_transform(X)
        y = scaler_y.fit_transform(y.reshape(-1, 1)).flatten()

        # Optimize SVM
        svm_optimized = optimize_svm_rbf(X, y, n_trials=100)
        model = svm_optimized['model']

        # Select 100 instances
        n_instances = 100
        indices = np.random.choice(X.shape[0], size=n_instances, replace=False)
        X_tbx = X[indices, :]

        # Create Excel file
        excel_file = os.path.join(results_dir, f"sv_d{d}_normalized.xlsx")
        workbook = Workbook()
        sheet_rbf = workbook.active
        sheet_rbf.title = "RBFExplainer"

        # Store RBFExplainer results and timing
        rbf_shapley_values = []
        rbf_times = []
        for i, x in enumerate(X_tbx):
            start_time = time.time()
            explainer = RBFLocalExplainer(model)
            shapley_values = explainer.explain(x)
            rbf_times.append(time.time() - start_time)
            rbf_shapley_values.append(shapley_values)
            sheet_rbf.append([i] + shapley_values)

        # Store RBFExplainer timing
        sheet_rbf.append(["Execution Time"] + rbf_times)

        # Compute Shapley regression values with varying number of samples
        max_samples = min(2**d, 100000)
        sample_sizes = np.logspace(np.log10(100), np.log10(max_samples), num=10, dtype=int)

        shapley_regression_times = []
        for sample_size in sample_sizes:
            sheet_name = f"samples_{sample_size}"
            sheet = workbook.create_sheet(title=sheet_name)

            shapley_regression_values = []
            shapley_regression_times_instance = []

            for i, x in enumerate(X_tbx):
                start_time = time.time()
                
                k_vectors = np.array(explainer.compute_kernel_vectors(X[model.support_, :], x))
                samples_weights = kernel_shap_sampling(d, sample_size)
                samples = np.array([t[0] for t in samples_weights])
                weights = np.array([t[1] for t in samples_weights])

                samples = np.vstack([samples, np.ones(samples[0].shape)]).astype(int)
                weights = np.hstack([weights, 1e20])

                y_values = compute_y(k_vectors, samples, model.dual_coef_.squeeze())

                shap_reg = Shapley_regression(np.array(samples, dtype='float'), y_values, weights)
                shap_vals = shap_reg.coef_

                shapley_regression_times_instance.append(time.time() - start_time)

                shapley_regression_values.append(shap_vals)
                sheet.append([i] + shap_vals.tolist())

            shapley_regression_times.append(shapley_regression_times_instance)
            # Store Shapley regression timing in the sheet
            sheet.append(["Execution Time"] + shapley_regression_times_instance)

        # Save workbook
        workbook.save(excel_file)



# Plot results
fig, axes = plt.subplots(1, 4, figsize=(15, 5))
axes = axes.flatten()
n_instances = 100

for idx, d in enumerate(ds):
    # Load results
    excel_file = os.path.join(results_dir, f"sv_d{d}.xlsx")
    workbook = load_workbook(excel_file)

    # Extract RBFExplainer results
    sheet_rbf = workbook["RBFExplainer"]
    rbf_shapley_values = np.array([row[1:d+1] for row in sheet_rbf.iter_rows(min_row=1, max_row=n_instances, values_only=True)])
    # Find the "Execution Time" row and extract its values
    for row in sheet_rbf.iter_rows(values_only=True):
        if row[0] == "Execution Time":
            rbf_times = np.array(row[1:n_instances+1])
            break

    # Compute delta_sv for each sample size
    delta_sv = []
    sample_sizes = []
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("samples_"):
            sample_size = int(sheet_name.split("_")[1])
            sample_sizes.append(sample_size)
        else:
            continue
        sheet_name = f"samples_{sample_size}"
        sheet = workbook[sheet_name]
        shapley_regression_values = np.array([row[1:d+1] for row in sheet.iter_rows(min_row=1, max_row=n_instances, values_only=True)])

        errors = np.linalg.norm(shapley_regression_values - rbf_shapley_values, axis=1, ord=1) #/ abs(np.sum(rbf_shapley_values, axis=1))
        delta_sv.append(errors)

    delta_sv = np.array(delta_sv)

    # Plot error bars and fill_between
    for i in range(delta_sv.shape[1]):
        axes[idx].plot(sample_sizes, delta_sv[:, i], alpha=0.5, label=f"Instance {i+1}" if i < 5 else None)

    # mean_errors = np.mean(delta_sv, axis=1)
    # std_errors = np.std(delta_sv, axis=1)
    # axes[idx].plot(sample_sizes, mean_errors, color="black", label="Mean Error", linewidth=2)
    # mean_errors = np.mean(delta_sv, axis=1)
    # std_errors = np.std(delta_sv, axis=1)
    # axes[idx].errorbar(sample_sizes, mean_errors, yerr=std_errors, label="Error")
    # axes[idx].fill_between(sample_sizes, mean_errors - std_errors, mean_errors + std_errors, alpha=0.2)
    
    axes[idx].set_xscale("log")
    axes[idx].set_title(f"d={d}")
    axes[idx].set_xlabel("Number of Samples")
    axes[idx].set_ylabel("Relative Error")
    axes[idx].legend()

plt.tight_layout()
plt.show()
plt.savefig(f"results/delta_shapley_samples.png", dpi=500, format='png', bbox_inches='tight')


# Boxplot execution times
fig, axes = plt.subplots(1, 4, figsize=(15, 5))
axes = axes.flatten()

for idx, d in enumerate(ds):
    # Load results
    excel_file = os.path.join(results_dir, f"sv_d{d}.xlsx")
    workbook = load_workbook(excel_file)

    # Extract RBFExplainer execution times
    sheet_rbf = workbook["RBFExplainer"]
    for row in sheet_rbf.iter_rows(values_only=True):
        if row[0] == "Execution Time":
            rbf_times = np.array(row[1:n_instances+1])
            break

    # Extract Shapley regression execution times
    shapley_regression_times = []
    labels = ["RBFExplainer"]
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("samples_"):
            sample_size = int(sheet_name.split("_")[1])
            labels.append(f"samples_{sample_size}")
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                if row[0] == "Execution Time":
                    shapley_regression_times.append(row[1:n_instances+1])
                    break

    # Combine all times
    all_times = [rbf_times] + shapley_regression_times

    # Plot boxplot
    axes[idx].boxplot(all_times, labels=labels, showfliers=False)
    axes[idx].set_title(f"d={d}")
    axes[idx].set_ylabel("Time (s)")
    axes[idx].tick_params(axis='x', rotation=45)

plt.tight_layout()
plt.show()
plt.savefig(f"results/time.png", dpi=500, format='png', bbox_inches='tight')


