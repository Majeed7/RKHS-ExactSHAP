import os
import pickle
import math
import numpy as np
import pandas as pd
from sklearn.svm import SVC, SVR
from sklearn.metrics import accuracy_score, mean_squared_error
from sklearn.model_selection import train_test_split, KFold
from openpyxl import load_workbook, Workbook
from sklearn.preprocessing import LabelEncoder
from sklearn.utils.multiclass import type_of_target
import sys
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.model_selection import GridSearchCV
from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier

from utils.real_datasets import load_dataset

from sklearn.gaussian_process import GaussianProcessRegressor, GaussianProcessClassifier
from sklearn.gaussian_process.kernels import RBF, ConstantKernel as C

import warnings
warnings.filterwarnings("ignore")

from sklearn.utils import shuffle

from sklearn.model_selection import GridSearchCV, train_test_split
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.metrics import accuracy_score, mean_absolute_percentage_error
import numpy as np

def train_rf(X, y, n_split=5):
    """
    Trains a Random Forest model using GridSearchCV and returns the performance score.
    
    Args:
    - X: Features data (input matrix)
    - y: Target labels
    
    Returns:
    - best_score: Best performance score (accuracy for classification, MSE for regression)
    """
    
    # Check if the problem is classification or regression
    problem_type = type_of_target(y)
    
    if problem_type == 'continuous' or problem_type == 'unknown':
        # If the target variable is numeric, treat it as regression
        model = RandomForestRegressor(random_state=42)
        scoring = 'neg_mean_squared_error'  # For regression, we use negative MSE as scoring
    else:
        # If the target variable is categorical, treat it as classification
        model = RandomForestClassifier(random_state=42)
        scoring = 'accuracy'  # For classification, use accuracy as scoring
    
    # Hyperparameter grid
    param_grid = {
        'n_estimators': [500, 1000],
        'max_depth': [None, 10, 20],
        'min_samples_split': [2, 5, 10]
    }
    
    # Split the data into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    
    # Initialize GridSearchCV with KFold cross-validation
    grid_search = GridSearchCV(estimator=model, param_grid=param_grid, cv=n_split, scoring=scoring, n_jobs=-1)
    
    # Fit the model using GridSearchCV
    grid_search.fit(X_train, y_train)
    
    # Get the best model and evaluate on the test set
    best_model = grid_search.best_estimator_
    
    # Make predictions and evaluate the model
    y_pred = best_model.predict(X_test)
    
    if isinstance(best_model, RandomForestClassifier):
        # If classification, use accuracy score
        best_score = accuracy_score(y_test, y_pred)
    else:
        # If regression, use mean squared error
        best_score = calculate_regression_scores(y_test, y_pred)["MAPE"] #mean_absolute_percentage_error(y_test, y_pred)
    
    # Return the best score
    return best_score, grid_search.cv_results_['mean_test_score'], grid_search.cv_results_['std_test_score']

def train_signle_gp(X, y, test_size=0.2):
    """
    Train a Gaussian Process on a train-test split.

    Parameters:
        X (array-like): Feature matrix of shape (n_samples, n_features).
        y (array-like): Target vector of shape (n_samples,).
        test_size (float): Proportion of data to include in the test set (default is 0.2).

    Returns:
        test_score (float): Performance score on the test set (MSE or Accuracy).
    """

    # Shuffle data before splitting
    X, y = shuffle(X, y, random_state=42)

    # Check the type of problem (classification or regression)
    problem_type = type_of_target(y)
    
    if problem_type == 'continuous' or problem_type == 'unknown':
        is_classification = False  # Regression problem
    else:
        is_classification = True   # Classification problem

    # Define the kernel for GP
    kernel = C(1.0, (1e-4, 1e1)) * RBF(1.0, (1e-4, 1e1))

    # Initialize the model
    if is_classification:
        gp = GaussianProcessClassifier(kernel=kernel, optimizer='fmin_l_bfgs_b')
    else:
        gp = GaussianProcessRegressor(kernel=kernel, optimizer='fmin_l_bfgs_b')

    # Split the data into training and test sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=42)

    # Train the GP on the training data
    gp.fit(X_train, y_train)

    # Make predictions on the test data
    y_pred = gp.predict(X_test)

    # Evaluate based on whether it's regression or classification
    if is_classification:
        # Calculate Accuracy for classification
        test_score = accuracy_score(y_test, y_pred)
        print(f"Test Accuracy: {test_score}")
    else:
        # Calculate Mean Squared Error for regression
        test_score = mean_squared_error(y_test, y_pred)
        print(f"Test MSE: {test_score}")

    return test_score

def train_gp(X, y, n_splits=5):
    """
    Train a Gaussian Process using k-fold cross-validation.
    
    Parameters:
        X (array-like): Feature matrix of shape (n_samples, n_features).
        y (array-like): Target vector of shape (n_samples,).
        n_splits (int): Number of folds for cross-validation (default is 5).
        
    Returns:
        avg_score (float): Average score (MSE or Accuracy) across folds.
        std_score (float): Standard deviation of the scores across folds.
    """
    
    # Check the type of problem (classification or regression) using sklearn utility
    problem_type = type_of_target(y)
    
    if problem_type == 'continuous' or problem_type == 'unknown':
        is_classification = False  # Regression problem
    else:
        is_classification = True   # Classification problem

    # Define the kernel for GP
    kernel = C(1.0, (1e-4, 1e1)) * RBF(1.0, (1e-4, 1e1))

    # Initialize the model
    if is_classification:
        gp = GaussianProcessClassifier(kernel=kernel, optimizer='fmin_l_bfgs_b', n_jobs=-1)
    else:
        gp = GaussianProcessRegressor(kernel=kernel, optimizer='fmin_l_bfgs_b', j_jobs=-1)

    # Setup K-fold Cross Validation
    kf = KFold(n_splits=n_splits, shuffle=True, random_state=42)

    # To store cross-validation results
    scores = []

    # Perform cross-validation
    for train_index, val_index in kf.split(X):
        X_train_fold, X_val_fold = X[train_index], X[val_index]
        y_train_fold, y_val_fold = y[train_index], y[val_index]
        
        # Train the GP on the training fold
        gp.fit(X_train_fold, y_train_fold)
        
        # Make predictions on the validation fold
        y_pred = gp.predict(X_val_fold)
        
        # Evaluate based on whether it's regression or classification
        if is_classification:
            # Calculate Accuracy for classification
            score = accuracy_score(y_val_fold, y_pred)
        else:
            # Calculate Mean Squared Error for regression
            score = mean_squared_error(y_val_fold, y_pred)
        
        scores.append(score)

    # Calculate average and standard deviation of scores
    avg_score = np.mean(scores)
    std_score = np.std(scores)

    # Print the results
    if is_classification:
        print(f"Average Accuracy across {n_splits}-folds: {avg_score}")
    else:
        print(f"Average MSE across {n_splits}-folds: {avg_score}")
        
    print(f"Standard Deviation of score: {std_score}")
    
    return avg_score, std_score

def reload_dataset(dataset_name):
    """Reload the dataset by its name."""
    return load_dataset(dataset_name)

def calculate_classification_scores(y_true, y_pred):
    """Calculate classification accuracy."""
    return {"Accuracy": accuracy_score(y_true, y_pred)}

def calculate_regression_scores(y_true, y_pred):
    """Calculate regression MSE."""
    return {"MSE": mean_squared_error(y_true, y_pred), "MAPE": mean_absolute_percentage_error(y_true, y_pred)}

def select_features_incrementally(X, y, ranked_features):
    """Select features incrementally and evaluate performance."""

    performance = []
    selected_features = []

    total_features = X.shape[1]
    i = math.ceil(total_features * 0.05)  # Start by selecting the top 10% of features

    while i <= (total_features / 2):
        # Select the top `i` ranked features
        selected_features = ranked_features[:i]

        # Subset the data for training and testing
        X_subset = X[:, selected_features]

        # Train the GP model on the selected features and evaluate
        best_score, avg_score, std_score = train_rf(X_subset, y) #  avg_score, std_score = train_gp(X_subset, y) # avg_score, std_score = train_single_gp(X_subset, y) # 

        # Track performance for this number of selected features
        performance.append({
            'num_features': i,
            'best_score': best_score,
            'avg_score': avg_score,
            'std_score': std_score
        })

        # Increment features by 10% of the total number of features
        i = math.ceil(total_features * 0.05 * (len(performance) + 1))  # Add 10% more each time

        # Ensure i does not exceed the total number of features
        if i > (total_features / 2):
            break #i = int(total_features / 2)

    return performance

def main():
    # Load the Excel file with feature importance data
    feature_importance_file = "results/fs_importance.xlsx"
    wb = load_workbook(feature_importance_file)

    # Create a new workbook for storing results
    results_wb = Workbook()

    sheetnames = wb.sheetnames

    # Process datasets in the Excel sheet
    for sheet_name in sheetnames:
        try:
            # if sheet_name in ['keggdirected']: 
            #     continue 
            print(f"Processing dataset: {sheet_name}")
            sheet = wb[sheet_name]

            # Reload the dataset
            X, y = reload_dataset(sheet_name)

            # Determine if it's classification or regression
            is_classification = type_of_target(y) in ["binary", "multiclass"]
            if is_classification:
                y = LabelEncoder().fit_transform(y)

            result_sheet = results_wb.create_sheet(title=sheet_name)

            # Set column titles dynamically
            score_titles = ["Accuracy"] if is_classification else ["MSE"]
            result_sheet.append(["Feature Selector"] + score_titles)

            # Process each feature selector (row) in the sheet
            for row in sheet.iter_rows(min_row=2, values_only=True):
                feature_selector = row[0]
                if feature_selector is None: break
                feature_importance = np.array(row[2:])

                # Use the ranking directly from the feature_importance array
                # Sorting by importance
                if row[0] == 'HSIC-SV':
                    ranked_features = np.argsort(-feature_importance)
                else:
                    ranked_features = np.argsort(-np.abs(feature_importance))

                # Incrementally evaluate performance with the selected features
                performance = select_features_incrementally(X, y, ranked_features)

                # First row for the best scores
                best_row = [feature_selector + "_best"] + [result['best_score'] for result in performance]
                result_sheet.append(best_row)

                # Second row for the average scores
                # avg_row = [feature_selector + "_avg"] + [result['avg_score'] for result in performance]
                # result_sheet.append(avg_row)

                # # Third row for the standard deviation scores
                # std_row = [feature_selector + "_std"] + [result['std_score'] for result in performance]
                # result_sheet.append(std_row)

            # Save the results to a new Excel file
            results_wb.save(f"results/incremental_feature_rf_bigestimators.xlsx")

        except Exception as e:
            print(f"{sheet_name} could not be processed! Error: {e}")
            continue


if __name__ == "__main__":
    main()
